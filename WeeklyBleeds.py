import streamlit as st 
import pandas as pd
import os
import gspread
import random
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import time
from pathlib import Path
from streamlit_gsheets import GSheetsConnection

st.set_page_config(
    page_title = 'VL TRACKER'
)
st.success('WELCOME, this app was developed by Dr. Luminsa Desire, for any concern, reach out to him at desireluminsa@gmail.com')
current_time = time.localtime()
week = time.strftime("%U", current_time)
week = int(week) + 1



st.markdown(f"******* **REMINDER!! we are currently in week {week}** *****")

st.markdown('**FIRST RENAME THESE COLUMNS BEFORE YOU PROCEED:**')
col1, col2, col3 = st.columns([1,1,1])
col1.markdown('Rename the **HIV CLINIC NO.** column to **A**')
col1.markdown('Rename the **ART START DATE** column to **AS**')
col2.markdown('Rename the **RETURN VISIT DATE** column to **RD**')
col2.markdown('Rename the **RETURN VISIT DATE_1** column to **RD1**')
col3.markdown('Rename the **HIV VIRAL LOAD DATE** column to **VD**')
col3.markdown('Rename the **HIV VIRAL LOAD DATE_1** column to **VD1**')

file = st.file_uploader(label='Upload your EMR extract here')

ext = None
if file is not None:
    # Get the file name
    fileN = file.name
    ext = os.path.basename(fileN).split('.')[1]
df = None
if file is not None:
    if ext !='xlsx':
        st.write('Unsupported file format, first save the excel as xlsx and try again')
        st.stop()
    else:
        df = pd.read_excel(file)
        st.write('Excel accepted')

    if df is not None:
        columns = ['A', 'AS', 'VD', 'VD1', 'RD', 'RD1']
        cols = df.columns.to_list()
        if not all(column in cols for column in columns):
            missing_columns = [column for column in columns if column not in cols]
            for column in missing_columns:
                st.markdown(f' **ERROR !!! {column} is not in the file uploaded**')
                st.markdown('**First rename all the columns as guided above**')
                st.stop()
        else:
            # Convert 'A' column to string and create 'ART' column with numeric part
            df['A'] = df['A'].astype(str)
            df['ART'] = df['A'].str.replace('[^0-9]', '', regex=True)
            df['ART'] = pd.to_numeric(df['ART'], errors= 'coerce')
            df = df[df['ART']>0]
            df.dropna(subset='ART', inplace=True)
            #st.write(df.shape[0])
            df = df.copy()
            #CONVERTING DATES TO STRINGS
            df[['AS', 'RD', 'VD','VD1','RD1']] = df[['AS', 'RD', 'VD', 'VD1','RD1']].astype(str)

            df['AS'] = df['AS'].str.replace('/', '*', regex=True)
            df['RD'] = df['RD'].str.replace('/', '*', regex=True)
            df['VD'] = df['VD'].str.replace('/', '*',regex=True)
            #df['LD'] = df['LD'].str.replace('/', '*',regex=True)
            df['RD1'] = df['RD1'].str.replace('/', '*',regex=True)
            df['VD1'] = df['VD1'].str.replace('/', '*',regex=True)

            df['AS'] = df['AS'].str.replace('-', '*',regex=True)
            df['RD'] = df['RD'].str.replace('-', '*',regex=True)
            df['VD'] = df['VD'].str.replace('-', '*',regex=True)
            #df['LD'] = df['LD'].str.replace('-', '*',regex=True)
            df['RD1'] = df['RD1'].str.replace('-', '*',regex=True)
            df['VD1'] = df['VD1'].str.replace('-', '*',regex=True)

            df['AS'] = df['AS'].str.replace('00:00:00', '')
            df['RD'] = df['RD'].str.replace('00:00:00', '')
            df['VD'] = df['VD'].str.replace('00:00:00', '')
            #df['LD'] = df['LD'].str.replace('00:00:00', '')
            df['RD1'] = df['RD1'].str.replace('00:00:00', '')
            df['VD1'] = df['VD1'].str.replace('00:00:00', '')

            df[['Ayear', 'Amonth', 'Aday']] = df['AS'].str.split('*', expand = True)
            df[['Ryear', 'Rmonth', 'Rday']] = df['RD'].str.split('*', expand = True)
            df[['Vyear', 'Vmonth', 'Vday']] = df['VD'].str.split('*', expand = True)
            #df[['Lyear', 'Lmonth', 'Lday']] = df['LD'].str.split('*', expand = True)
            df[['RD1year', 'RD1month', 'RD1day']] = df['RD1'].str.split('*', expand = True)
            df[['VD1year', 'VD1month', 'VD1day']] = df['VD1'].str.split('*', expand = True)

            df['AS'] = df['AS'].str.replace('*', '/',regex=True)
            df['RD'] = df['RD'].str.replace('*', '/',regex=True)
            df['VD'] = df['VD'].str.replace('*', '/',regex=True)
            #df['LD'] = df['LD'].str.replace('*', '/',regex=True)
            df['RD1'] = df['RD1'].str.replace('*', '/',regex=True)
            df['VD1'] = df['VD1'].str.replace('*', '/',regex=True)


            #SORTING THE VIRAL LOAD YEARS
            df[['Vyear', 'Vmonth', 'Vday']] =df[['Vyear', 'Vmonth', 'Vday']].apply(pd.to_numeric, errors = 'coerce') 
            df['Vyear'] = df['Vyear'].fillna(2022)
            a = df[df['Vyear']>31].copy()
            b = df[df['Vyear']<32].copy()
            b = b.rename(columns={'Vyear': 'Vday1', 'Vday': 'Vyear'})
            b = b.rename(columns={'Vday1': 'Vday'})
            df = pd.concat([a,b])
            dfa = df.shape[0]

            #SORTING THE VIRAL LOAD DATE1 YEARS
            df[['VD1year', 'VD1month', 'VD1day']] =df[['VD1year', 'VD1month', 'VD1day']].apply(pd.to_numeric, errors = 'coerce')
            df['VD1year'] = df['VD1year'].fillna(2022)
            a = df[df['VD1year']>31].copy()
            b = df[df['VD1year']<32].copy()
            b = b.rename(columns={'VD1year': 'VD1day1', 'VD1day': 'VD1year'})
            b = b.rename(columns={'VD1day1': 'VD1day'})
            df = pd.concat([a,b])
            dfb = df.shape[0]
            #SORTING THE RETURN VISIT DATE YEARS
            df[['Rday', 'Ryear']] = df[['Rday', 'Ryear']].apply(pd.to_numeric, errors='coerce')
            df['Ryear'] = df['Ryear'].fillna(2022)
            a = df[df['Ryear']>31].copy()
            b = df[df['Ryear']<32].copy()
            b = b.rename(columns={'Ryear': 'Rday1', 'Rday': 'Ryear'})
            b = b.rename(columns={'Rday1': 'Rday'})

            df = pd.concat([a,b])
            dfc = df.shape[0]
            #SORTING THE RETURN VISIT DATE1 YEAR
            df[['RD1day', 'RD1year']] = df[['RD1day', 'RD1year']].apply(pd.to_numeric, errors='coerce')
            df['RD1year'] = df['RD1year'].fillna(2022)
            a = df[df['RD1year']>31].copy()
            b = df[df['RD1year']<32].copy()
            b = b.rename(columns={'RD1year': 'RD1day1', 'RD1day': 'RD1year'})
            b = b.rename(columns={'RD1day1': 'RD1day'})
            df = pd.concat([a,b])

            dfd = df.shape[0]
            #SORTING THE ART START YEARS
            df[['Ayear', 'Amonth', 'Aday']] =df[['Ayear', 'Amonth', 'Aday']].apply(pd.to_numeric, errors = 'coerce')
            df['Ayear'] = df['Ayear'].fillna(2022)
            a = df[df['Ayear']>31].copy()
            b = df[df['Ayear']<32].copy()
            b = b.rename(columns={'Ayear': 'Aday1', 'Aday': 'Ayear'})
            b = b.rename(columns={'Aday1': 'Aday'})
            df = pd.concat([a,b])
            dfe = df.shape[0]
            #print(dfa, dfb, dfc, dfd, dfe)
            # AVOIDING NUMPY
            def fill(r):
                if r > 0:
                    return 'BLED'
                else:
                    return 'BLED'

            #SORTING THE TX_NEW

            def Eligible(x):
                if x <2024:
                    return ('ELLIGIBLE')
                else:
                    return('NOT')
             #LAST WEEK, THIS WEEK
            def this(a):
                if a == week-1:
                    return 'LAST WEEK'
                elif a == week:
                    return 'THIS WEEK'
                elif a == week+1:
                    return 'NEXT WEEK'
                else:
                    return a

            #FOR THOSE DUE
            def Due(x, y):
                if x ==2024:
                    if y < 4:
                        return 'ALREADY'
                    elif y > 3:
                        return 'BLED'     
                elif x == 2023:
                    if y > 6:
                        return 'ALREADY'
                    else:
                        return 'DUE'
                else:
                    return('DUE')

            #FOR WEEKS
            def week (c,a,b):
                if c ==2024:
                    if a == 4:
                        if 1 <= b <=7:
                            return 14
                        elif 8<= b <= 14:
                            return 15
                        elif 15 <= b <=21:
                            return 16
                        elif 22<= b <= 28:
                            return 17
                        elif 29 <= b <=30:
                            return 18
                    elif a == 5:
                        if 1 <= b <=5:
                            return 18
                        elif 6<= b <= 12:
                            return 19
                        elif 13 <= b <= 19:
                            return 20
                        elif 20 <= b <=26:
                            return 21
                        elif 27 <= b <= 31:
                            return 22
                    elif a == 6:
                        if 1 <= b <= 2:
                            return 22
                        elif 3 <= b <= 9:
                            return 23
                        elif 10 <= b <= 16:
                            return 24
                        elif 17 <= b <= 23:
                            return 25
                        elif 24 <= b <= 30:
                            return 26
                    else:
                        return None
                else:
                    return None
                    
            #APPLYING THE WEEK FORMULA ON VL DATES
            df[['Vday', 'Vyear','Vmonth']] = df[['Vday', 'Vyear', 'Vmonth']].apply(pd.to_numeric, errors='coerce')
            df['BWEEK'] = df.apply(lambda wee: week(wee['Vyear'],wee['Vmonth'], wee['Vday']), axis=1)
            #APPLYIN THE DUE FORMULA ON VL DATES
            df['DUE'] = df.apply(lambda due: Due(due['Vyear'], due['Vmonth']), axis=1)
            
            #APPLYING THE DUE FORMUA TO VD1 DATES
            df[['VD1day', 'VD1year','VD1month']] = df[['VD1day', 'VD1year', 'VD1month']].apply(pd.to_numeric, errors='coerce')
            df['DUE1'] = df.apply(lambda due: Due(due['VD1year'], due['VD1month']), axis=1)
            
            #APPLYING THE WEEK FORMULA TO RETURN VISIT DATES
            df[['Rday', 'Ryear','Rmonth']] = df[['Rday', 'Ryear', 'Rmonth']].apply(pd.to_numeric, errors='coerce')
            df['WEEK'] = df.apply(lambda row: week(row['Ryear'], row['Rmonth'], row['Rday']),axis=1)
            
            #APPLYING THE WEEK FORMULA TO RETURN VISIT DATES 1
            df[['RD1day', 'RD1year','RD1month']] = df[['RD1day', 'RD1year', 'RD1month']].apply(pd.to_numeric, errors='coerce')
            df['WEEK1'] = df.apply(lambda row: week(row['RD1year'], row['RD1month'], row['RD1day']),axis=1)
            #APPLYING THE FORMULA TO RULE OUT TX NEW
            df['Ayear'] = pd.to_numeric(df['Ayear'], errors= 'coerce')
            
            #APPLYING THE FORMULA TO RULE OUT TX NEW
            df['Ayear'] = pd.to_numeric(df['Ayear'], errors= 'coerce')
            df['ELL'] = df['Ayear'].apply(Eligible)
            
            #RECONSTRUCTING RETURN VISIT DATES
            df['Ryear'] = pd.to_numeric(df['Ryear'], errors= 'coerce')
            dfa = df[df['Ryear']==2024].copy()
            df['RD1year'] = pd.to_numeric(df['RD1year'], errors= 'coerce')
            dfb = df[df['RD1year']==2024].copy()

            dfa[['Rday', 'Ryear','Rmonth']] = dfa[['Rday', 'Ryear', 'Rmonth']].apply(pd.to_numeric, errors='coerce')
            CURRa = dfa[((dfa['Rmonth']>3) | ((dfa['Rmonth']==3) & (dfa['Rday'] >3)))].copy()
            dfb[['RD1day', 'RD1year','RD1month']] = dfb[['RD1day', 'RD1year', 'RD1month']].apply(pd.to_numeric, errors='coerce')
            CURRb = dfb[((dfb['RD1month']>3) | ((dfb['RD1month']==3) & (dfb['RD1day'] >3)))].copy()
            CURRb = CURRb.drop(columns=['WEEK'])
            CURRa = CURRa.drop(columns=['WEEK1'])
            CURRb = CURRb.rename(columns={'WEEK1': 'WEEK'})
            #..dfcurr = pd.concat([CURRa, CURRb])
            #..dfcurr['ELL'] = dfcurr['ELL'].astype(str)
            #..dfcurr = dfcurr[dfcurr['ELL'] == 'ELLIGIBLE'].copy()
            
            #COMPUTING WEEKLY BLEEDS
            #CHOOSE BLEEDS DONE IN THE QUARTER
            df[['Vday', 'Vyear','Vmonth']] = df[['Vday', 'Vyear', 'Vmonth']].apply(pd.to_numeric, errors='coerce')
            dfv = df[(df['Vyear']==2024) & (df['Vmonth']>3)].copy()
            dfv['DUE'] = dfv['DUE'].astype(str)
            dfv['DUE1'] = dfv['DUE1'].astype(str)
            
            #REBLEEDS
            REBLED = dfv[((dfv['DUE'] == 'BLED')& (dfv['DUE1'] == 'ALREADY'))].copy()
            #REBLED['STATUS'] = REBLED.apply(lambda d: fill(d['Rmonth']), axis=1)
            #REBLED['STATUS'] = REBLED['STATUS'].replace('BLED', 'REBLED')
            REBLED['STATUS'] = np.nan
            REBLED['STATUS'] = REBLED['STATUS'].fillna('REBLED')
            REBLEDpivo = pd.pivot_table(REBLED, index= 'BWEEK', values='A', aggfunc = 'count')
            REBLEDF = REBLEDpivo.reset_index()
            REBLEDF = REBLEDF.rename(columns={'A': 'REBLED'})
            
            #BLEEDS, original dataframe should be dfv
            BLED = dfv[((dfv['DUE'] == 'BLED')& (dfv['DUE1'] == 'DUE'))].copy()
            #BLED['STATUS'] = BLED.apply(lambda d: fill(d['Rmonth']), axis=1)
            BLED['STATUS'] = np.nan
            BLED['STATUS'] = BLED['STATUS'].fillna('BLED')
            BLEDpivo = pd.pivot_table(BLED, index= 'BWEEK', values='A', aggfunc = 'count')
            BLEDF = BLEDpivo.reset_index()
            BLEDF = BLEDF.rename(columns={'A': 'BLED'})
            BLEDF = BLEDF.rename(columns={'A': 'BLED'})
            
            #TOTAL BLEEDS
            TOTAL = pd.pivot_table(dfv, index= 'BWEEK', values='A', aggfunc = 'count')
            TOTAL = TOTAL.reset_index()   
            TOTAL = TOTAL.rename(columns={'A':'TOTAL BLEEDS'}) 
            
            #####FIRST PIVOT FROM BLED, REBLED, AND TOTAL BLEEDS   
            BLEDF['BWEEK'] = pd.to_numeric(BLEDF['BWEEK'], errors='coerce')     
            REBLEDF['BWEEK'] = pd.to_numeric(REBLEDF['BWEEK'], errors='coerce')
            dfy = pd.merge(BLEDF, REBLEDF, on = 'BWEEK', how ='outer')
            TOTAL['BWEEK'] = pd.to_numeric(TOTAL['BWEEK'], errors='coerce')
            dfy['BWEEK'] = pd.to_numeric(dfy['BWEEK'], errors='coerce')
            dfg = pd.merge(dfy, TOTAL, on = 'BWEEK', how='outer')
            dfg = dfg.rename(columns = {'BWEEK': 'WEEK'})
            efg = dfg.copy()
            current_time = time.localtime()
            week = time.strftime("%U", current_time)
            week = int(week) + 1
            w = week-1
            we = efg[efg['WEEK']==w].copy()
            test = we.shape[0]
            if test==0:
                we = 0
            else:
                try:
                  we = int(we.iloc[0,1])
                except:
                    we = 0
            dfg = dfg.set_index('WEEK')
            weekly = dfg.copy()
           
             #DETERMINING TX_CURR
            #MODIFY HERE TO INCLUDE 2025 LATER
            CURR = df.copy()
            CURR[['Rday', 'Ryear','Rmonth']] = CURR[['Rday', 'Ryear', 'Rmonth']].apply(pd.to_numeric, errors='coerce')
            CURR = CURR[CURR['Ryear']==2024].copy()
            CURR = CURR[((CURR['Rmonth']>3) | ((CURR['Rmonth']==3) & (CURR['Rday'] >3)))].copy()
            CUR = CURR.copy()
            APPONT = CURR.copy()
            
            #NEW CODE..................
            #MISSED APPOINTMENT BUT DUE FOR VL
            CURRa = CURRa[CURRa['ELL'] == 'ELLIGIBLE'].copy()
            CURRa = CURRa[['A', 'RD','Rmonth', 'Rday', 'AS', 'VD', 'DUE', 'WEEK']]
            CURRa['Rmonth'] = pd.to_numeric(CURRa['Rmonth'], errors = 'coerce')
            MARCHm = CURRa[CURRa['Rmonth']==3].copy()
            CURRm =  CURRa[CURRa['Rmonth'].isin([4,5,6])].copy()
            CURRm['WEEK'] = pd.to_numeric(CURRm['WEEK'], errors='coerce')
            current_time = time.localtime()
            week = time.strftime("%U", current_time)
            week = int(week) + 1
            CURRm = CURRm[CURRm['WEEK'] < week].copy()
            CURR['DUE'] = CURR['DUE'].astype(str)
            CURRm = CURRm[CURRm['DUE'] =='DUE'].copy()
            ram = CURRm.shape[0]
            pivotm = pd.pivot_table(CURRm, index = 'WEEK', values = 'A', aggfunc = 'count')
            pivotm = pivotm.reset_index()
            pivotm = pivotm.set_index('WEEK')
            MARCHm = MARCHm[MARCHm['DUE']=='DUE'].copy()
            MARCHm['WEEK'] = MARCHm['WEEK'].fillna('MARCH')
            rbm = MARCHm.shape[0]
            pivotrbm = pd.pivot_table(MARCHm, index = 'WEEK', values = 'A', aggfunc = 'count')
            pivotrbm = pivotrbm.reset_index()
            pivotrbm = pivotrbm.set_index('WEEK')
            pivotmissed = pd.concat([pivotrbm, pivotm])
            pivotmissed = pivotmissed.rename(columns= {'A': 'MISSED, DUE FOR VL'})
            #pivotmissed['WEEK'] =  pd.to_numeric(pivotmissed['WEEK'], errors='coerce')
            MISSED = pd.concat([MARCHm, CURRm])
            rm = MISSED.shape[0]

            #RETURNED BUT NOT BLED
            CURRb = CURRb[CURRb['ELL'] == 'ELLIGIBLE'].copy()
            CURRb = CURRb[['A', 'RD', 'RD1','RD1month', 'RD1day', 'AS', 'VD', 'DUE', 'WEEK']]
            CURRb['RD1month'] = pd.to_numeric(CURRb['RD1month'], errors = 'coerce')
            MARCH = CURRb[CURRb['RD1month']==3].copy()
            CURRr =  CURRb[CURRb['RD1month'].isin([4,5,6])].copy()
            CURRr = CURRr[CURRr['WEEK'] < week].copy()
            CURRr = CURRr[CURRr['DUE'] =='DUE'].copy()
            ra = CURRr.shape[0]
            pivotr = pd.pivot_table(CURRr, index = 'WEEK', values = 'A', aggfunc = 'count')
            MARCH = MARCH[MARCH['DUE']=='DUE']
            rb = MARCH.shape[0]
            MARCH['WEEK'] = MARCH['WEEK'].fillna('MARCH')
            pivotrb = pd.pivot_table(MARCH, index = 'WEEK', values = 'A', aggfunc = 'count')
            pivotreturned = pd.concat([pivotrb, pivotr])
            pivotreturned = pivotreturned.reset_index()
            pivotreturned = pivotreturned.rename(columns= {'A': 'RETURNED NOT BLED'})
            pivotreturned['WEEK.'] = pivotreturned.apply(lambda q: this(q['WEEK']), axis=1)
            pivotreturned = pivotreturned.drop(columns='WEEK')
            pivotreturned = pivotreturned.set_index('WEEK.')
            RETURNED = pd.concat([MARCH, CURRr])
            r = RETURNED.shape[0]

            #CLIENTS ON APPT VS ELIGIBLE
            APPONT['Rmonth'] = pd.to_numeric(APPONT['Rmonth'], errors = 'coerce')
            APPONT = APPONT[APPONT['Rmonth'].isin([4,5,6])].copy()
            APPONT['WEEK'] = pd.to_numeric(APPONT['WEEK'], errors = 'coerce')
            APPONT = APPONT[APPONT['WEEK']>=week].copy()
            pivoappt = pd.pivot_table(APPONT, index = 'WEEK', values = 'A', aggfunc = 'count')
            pivoappt = pivoappt.reset_index()
            pivoappt = pivoappt.rename(columns={'A': 'ON APPT'})
            APPONT = APPONT[APPONT['ELL'] == 'ELLIGIBLE'].copy()
            APPONT = APPONT[APPONT['DUE']=='DUE'].copy()
            pivoapptd = pd.pivot_table(APPONT, index = 'WEEK', values = 'A', aggfunc = 'count')
            pivoapptd = pivoapptd.reset_index()
            pivoapptd = pivoapptd.rename(columns={'A': 'DUE FOR BLEEDING'})
            appt = pd.merge(pivoappt, pivoapptd, on = 'WEEK', how='outer')
            appt['WEEK'] = pd.to_numeric(appt['WEEK'], errors = 'coerce')
            appt['WEEK.'] = appt.apply(lambda q: this(q['WEEK']),axis = 1)
            next = week+1
            ee = appt[appt['WEEK'] == next]
            tes = ee.shape[0]
            if tes==0:
                el = 0
            else:
                try:
                  el = int(ee.iloc[0,2])
                except:
                    el = 0
            appt = appt.set_index('WEEK.')
            appt = appt.drop(columns = 'WEEK')
            
        
            #DETERMINING TX_CURR
            #MODIFY HERE TO INCLUDE 2025 LATER
            # CURR = df.copy()
            # CURR[['Rday', 'Ryear','Rmonth']] = CURR[['Rday', 'Ryear', 'Rmonth']].apply(pd.to_numeric, errors='coerce')
            # CURR = CURR[CURR['Ryear']==2024].copy()
            # CURR = CURR[((CURR['Rmonth']>3) | ((CURR['Rmonth']==3) & (CURR['Rday'] >3)))].copy()
            # CUR = CURR.copy()
            a = CUR.shape[0]
            CUR['ELL'] = CUR['ELL'].astype(str)
            #COUNT THOSE NOT ELLIGIBLE
            f = CUR[CUR['ELL'] =='NOT'].shape[0]
            CURB = CUR[CUR['ELL'] =='ELLIGIBLE'].copy()
            b = CURB[CURB['DUE']== 'ALREADY'].shape[0]
            c = CURB[CURB['DUE']=='BLED'].shape[0]
            d = CURB[CURB['DUE']=='DUE'].shape[0]
            E =  b + c + f 
            G = int((E/a)*100)
            H = int((a*0.95)- E)
            data = {'TX_CURR' : [a],
            'NOT DUE' : [E],
            'VL COV' : [G],
            'BALANCE TO 95%' : [H],
            'DUE FOR VL' : [d]}
            PERFORMANCE = pd.DataFrame(data)
            PERFORMANCE = PERFORMANCE.set_index('TX_CURR')
            #st.markdown(f'**NOTE!! This EMR shows {d} that are not yet bled**')
            CURB = CURB.rename(columns = {'DUE' : 'VL STATUS'})
            linelist = CURB[['A', 'AS', 'RD', 'Ryear', 'Rmonth', 'Rday', 'VD', 'VL STATUS']].copy()
            linelist['AS'] = linelist['AS'].astype(str)
            linelist['AS'] = linelist['AS'].str.replace('*', '/')
            linelist['AS'] = linelist['AS'].str.replace('NaT', '')
            linelist['RD'] = linelist['RD'].astype(str)
            linelist['RD'] = linelist['RD'].str.replace('*', '/')
            linelist['RD'] = linelist['RD'].str.replace('NaT', '')
            linelist['VD'] = linelist['VD'].astype(str)
            linelist['VD'] = linelist['VD'].str.replace('*', '/')
            linelist['VD'] = linelist['VD'].str.replace('NaT', '')
            linelist['VL STATUS'] = linelist['VL STATUS'].astype(str)
            linelist = linelist[linelist['VL STATUS']== 'DUE']
            linelist = linelist.rename(columns = {'A': 'ART-NO', 'RD': 'RETURN DATE', 'VD': 'VIRAL LOAD DATE', 'AS':'ART START'})
            linelist[['Ryear', 'Rmonth', 'Rday']] = linelist[['Ryear', 'Rmonth', 'Rday']].apply(pd.to_numeric, errors='coerce')
            linelist = linelist.sort_values(by = ['Ryear', 'Rmonth', 'Rday'])
            #
            #DUE FOR VL PER MONTH
            bymonth = pd.pivot_table(linelist, index='Rmonth', values= 'ART-NO', aggfunc='count')
            bymonth = bymonth.reset_index()
            bymonth = bymonth.rename(columns={'ART-NO': 'DUE PER MONTH', 'Rmonth':'MONTH'})
            bymonth['MONTH'] = bymonth['MONTH'].astype(int)
            bymonth = bymonth.set_index('MONTH')
            bymonth = bymonth.transpose()

            #DISPLAYS
            st.markdown(f'**NOTE!! This EMR shows {d} that are not yet bled**')
            weeks = 26+1 - week
            bleed = int(H/weeks)
            st.markdown(f'**You will have to bleed {bleed} clients per week in the remaining {weeks} weeks if you are to hit 95%**')
            st.markdown("**_THIS TX_CURR**, is from March 4th, (doesn't exclude TIs, TOs, and TX_news)_")
            st.write(PERFORMANCE)
            st.write('**No. Not BLED IN EACH MONTH ON APPOINTMENT**')
            st.table(bymonth)
            weekly = weekly.rename(columns={'A':'REBLED'})
            cola, colb = st.columns([1,1])
            cola.write('**BLEEDS  DONE PER WEEK**')
            cola.write(weekly)
            colb.write('**No. ELLIGIBLE FOR BLEEEDING IN THE COMING WEEKS**')
            colb.write(appt)
            st.markdown(f'**This emr shows {r} clients that returned and were not bled, {rb} in March and {ra} this quarter, scroll down to download this list and audit it first**')
            st.markdown(f'**Also there are {rm} cients that have missed appointment but are due for VL; {rbm} in March and {ram} this quarter, find them in the VL LINELIST**')
            cole, colf = st.columns([1,1])
            cole.markdown('**RETURNED, NOT BLED**')
            cole.write(pivotreturned)
            colf.markdown('**MISSED, DUE FOR VL**')
            colf.write(pivotmissed)
            st.markdown('**Sample linelist**')
            st.write(linelist.head(5))
            #st.write(appt)
    

if df is not None:
    def download_weekly(df):
        st.write(f"<h6>DOWNLOAD WEEKLY BLEEDS AND NEW VL LINELIST</h6>", unsafe_allow_html=True)

        if df is not None:
            dft = weekly.copy()
            csv_data = dft.to_csv(index=True)

                    # Create a download button for each facility

            st.download_button(
                        label=" DOWNLOAD WEEKLY BLEEDS",
                        data=csv_data,
                        file_name=f"WEEKLY_BLEEDS.csv",
                        mime="text/csv"
                    )


    def main():
        # Call the download functions
        download_weekly(df)


    if __name__ == "__main__":
        main()

if df is not None:
    def download_returned(df):
        st.write(f"<h6>DOWNLOAD CLIENTS THAT RETURNED BUT WERE NOT BLED</h6>", unsafe_allow_html=True)

        if df is not None:
            dft = RETURNED.copy()
            dft = dft[[ 'A', 'RD', 'RD1','RD1month', 'RD1day', 'AS', 'VD', 'DUE']]
            dft['VD'] = dft['VD'].replace('NaT', '')
            dft = dft.rename(columns = {'RD': 'RETURN DATE', 'RD1': 'RETURN DATE1','A': 'ART-NO.', 'DUE': 'VL STATUS', 'VD': 'VIRAL LOAD DATE'})
            dft[['RD1month', 'RD1day']] = dft[['RD1month', 'RD1day']].apply(pd.to_numeric, errors = 'coerce')
            dft = dft.sort_values(by = ['RD1month', 'RD1day'])
            
            csv_data = dft.to_csv(index=False)

                    # Create a download button for each facility

            st.download_button(key ='XXX',
                        label=" DOWNLOAD CLIENTS THAT RETURNED BUT WERE NOT BLED",
                        data=csv_data,
                        file_name=f"RETURNED_NOT_BLED.csv",
                        mime="text/csv"
                    )
    def main():
        # Call the download functions
        download_returned(df)
    if __name__ == "__main__":
        main()

if df is not None:
        ran = random.random()
        rand = round(ran,2)
        #st.success(f'Your file will be downloaded as "VL LINELIST {rand}.xlsx" in your Downloads folder.')
    #if st.button('DOWNLOAD CURRENT LINELIST'):
    #if st.button('DOWNLOAD CURRENT LINELIST'):
        wb = Workbook()
        ws = wb.active
        # Convert DataFrame to Excel
        for r_idx, row in enumerate(linelist.iterrows(), start=1):
            for c_idx, value in enumerate(row[1], start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        ws.insert_rows(0)

        blue = PatternFill(fill_type = 'solid', start_color = 'C8CDCD')
        # ws.column_dimensions['H'].width = 14

        for num in range (1, ws.max_row+1):
            for letter in ['G', 'H']:
                ws[f'{letter}{num}'].font = Font(b= True, i = True)
                ws[f'{letter}{num}'].font = Font(b= True, i = True)
                ws[f'{letter}{num}'].fill = blue
                ws[f'{letter}{num}'].border = Border(top = Side(style = 'thin', color ='000000'),
                                                    right = Side(style = 'thin', color ='000000'),
                                                    left = Side(style = 'thin', color ='000000'),
                                                    bottom = Side(style = 'thin', color ='000000'))
        ws['A1'] ='ART NO.'
        ws['B1'] = 'ART START DATE'
        ws['C1'] = 'RETURN VISIT DATE'
        ws['D1'] = 'Ryear'
        ws['E1'] = 'Rmonth'
        ws['F1'] = 'Rday'
        ws['G1'] = 'VIRAL LOAD DATE' 
        ws['H1'] = 'VL STATUS'
        ws['I1']  = 'WAS CLIENT BLED (Y/N)'
        ws['J1'] = 'IF YES, HAS EMR BEEN UPDATED (Y/N)'


        letters = ['B', 'C', 'G']
        for letter in letters:
            ws.column_dimensions[letter].width =17

        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws['I1'].alignment = Alignment(wrap_text=True)
        ws.column_dimensions['J'].width = 20
        ws['J1'].alignment = Alignment(wrap_text=True)

        file_path = os.path.join(os.path.expanduser('~'), 'Downloads', f'VL LINELIST {rand}.xlsx')
        directory = os.path.dirname(file_path)
        Path(directory).mkdir(parents=True, exist_ok=True)

      # Save the workbook
        wb.save(file_path)
    
    # Serve the file for download
        with open(file_path, 'rb') as f:
            file_contents = f.read()
        
        st.download_button(label='Download VL LINELIST', data=file_contents,file_name=f'VL LINELIST {rand}.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
             
                        #SUBMISSION
conn = st.connection('gsheets', type=GSheetsConnection)
exist = conn.read(worksheet ='VL', usecols = list(range(12)), ttl=5)
existing = exist.dropna(how='all')

if df is not None:
     if 'ent' not in st.session_state:
         st.session_state.ent = ''

#     # Create two equal columns
     cola, colb = st.columns([1, 1])

#     # Text input in the first column
     st.session_state.ent = cola.text_input('Enter name of the facility here and press Enter', 
                                         value=st.session_state.ent, 
                                         placeholder="e.g Mateete HCIII")

#     # Button in the second column
     submit = colb.button('Submit') 
  
     Facility = st.session_state.ent  
     WEEK = week-1
     TX_CURR = a
     NO_WITH_VL = E
     VL_COV = G
     BALANCE_TO_95 =H
     TOTAL_DUE_FOR_VL = d
     No_OF_BLEEDS_IN_THE_WEEK = we
     ON_APPT_NEXT_WEEK_DUE = el
     ADJUSTED_WEEKLY_TARGET = bleed
     No_RETURNED_NOT_BLED = r
     MISSED_APPT_BUT_DUE = rm

     if submit:
          updated = pd.DataFrame({'FACILITY': [Facility],
                                    'WEEK':[B],
                                    'TX_CURR':[a],
                                    'NO_WITH_VL':[E],
                                    'VL COV': [G],
                                    'BALANCE_TO 95':[H],
                                    'TOTAL UNMET IN THIS EMR EXTRACT': [d],
                                    'No. OF BLEEDS IN THE WEEK' : [we],
                                    'No.ON APPT NEXT WEEK THAT ARE DUE FOR VL': [el],
                                    'ADJUSTED WEEKLY TARGET' : [bleed],
                                    'No. RETURNED BUT NOT BLED' : [r],
                                    'MISSED APPT BUT DUE FOR VL' : [rm]
                                })
          st.write(updated)
     # updated = pd.concat([existing, updated], ignore_index =True)
     # conn.update(worksheet = 'VL', data = updated)
     # st.success('Your data has been submitted')
       




    























 
