import streamlit as st 
import pandas as pd
import os
import random
#import numpy as np
from openpyxl import Workbook
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
#from openpyxl import * #load_workbook
#from openpyxl.styles import *
# from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


SEMBABULE = {'Ssembabule HC IV':2232,'Kyabi HC III':511,'Ntuusi HC IV':1006, 'Lwemiyaga HC III':959,
            'Makoole HC II':245,'Mateete HC III':2338, 'Lwebitakuli Gvt HC III':567,'Ntete HC II':78,'Sembabule Kabaale HC II':59}

BUKOMANSIMBI = {'Butenga HC IV':1434,'Mirambi HC III':327,'Kagoggo HC II': 88,'Kisojjo HC II GOVT':73,'Bigasa HC III':807,
            'Kitanda HC III':321,"St. Mary'S Maternity Home HC III": 79,'Kingangazzi HC II':205
              }
KALUNGU = {'Lukaya Health Care Center-Uganda Cares HC II': 3623, 
           'Bukulula HC IV': 1177, 'Kalungu Kabaale HC II GOVT': 107,'Kalungu HC III': 785,'Kalungu Kasambya HC III GOVT':364,
           'Kiragga HC III':173,'Kiti HC III':176,'Kyamulibwa Gvt HC III':404,'Lukaya HC III':681,'MRC Kyamulibwa HC II':481}

LYANTONDE ={'Kabatema HC II':117,
           'Kabayanda HC II':91,'Kaliiro HC III':473,'Kasagama HC III':459,
           'Kinuuka  HC III':317,'Lyakajura HC II': 384,'Lyantonde Hospital':4129,'Mpumudde HC III':460}


MASAKA_CITY ={'Bukoto HC III':451,
             'Kiyumba HC IV':854,'Masaka Police HC II':400,'Mpugwe HC III':316,'Nyendo HC II':337,'TASO Masaka CLINIC':8170}

MASAKA_DISTRICT ={'Bukakata HC III':641,'Bukeeri HC III':355,'Masaka Buwunga HC III GOVT':340,'Kamulegu HC III':590,'Kyannamukaaka HC IV':1345}

MPIGI ={'Bunjako HC III':502,'Buwama HC III':2115,'Mpigi   HC IV':3107,'Bujuuko HC III':336,
       'Sekiwunga HC III':332,'Nsamu-Kyali HC III':187,'Butoolo HC III':819,'Buyiga HC II':152,'Kampiringisa HC III':260,
       'Ggolo HC III':322,'Kituntu HC III':313,'Dona Medical Centre HC III':151,'Nindye HC III':261,'Muduuma HC III':762,
       'Nabyewanga HC II':134,'Bukasa HC II':59,'Fiduga HC III':22,'Kiringente Epi Centre HC II':78,'St. Elizabeth Kibanga Ihu HC III':35}

BUTAMBALA ={'Bulo HC III': 689,
           'Butambala Epi Centre HC III GOVT':212,'Gombe HOSPITAL': 3277,'Kitimba HC III': 202,'Kyabadaza HC III':387,'Ngando HC III':404
          }

KYOTERA ={'Kabira HC III GOVT':1171,
          'Kabuwoko Gvt HC III':237,'Kakuuto HC IV':2669,'Kalisizo Hospital':4029,'Kasasa HC III': 868,'Kasaali HC III': 1220,
          'Kasensero HC II':1304,'Kayanja HC II Lwankoni GOVT':107,'Kirumba  HC III':310,'Kyebe HC III':676,'Lwankoni HC III':260,
          'Mutukula HC III':548,'Mitukula HC III':1091,'Nabigasa HC III':860,'Rakai Health Sciences Program CLINIC':2647}

RAKAI = {'Buyamba HC III':892,
         'Byakabanda HC III':181,'Kacheera HC III':478,'Kibaale HC II GOVT':547,'Kibanda HC III':292,'Kimuli HC III':567,'Kifamba HC III':326,
         'Kyalulangira HC III':365,'Lwamaggwa Gvt HC III':980,'Lwanda HC III':750,'Rakai Hospital':3075,'Rakai Kiziba HC II GOVT':387}

GOMBA= {'Buyanja  HC II (Gomba)': 110,'Gomba Kanoni HC III GOVT': 1289,'Kifampa HC III': 860,'Kisozi HC III GOVT':390,
        'Kyai HC III': 364,'Maddu HC IV': 1999,'Mamba HC II':304,'Mpenja HC III': 405,'Ngomanene HC II': 99}

WAKISO= { 'Bulondo HC III':308,'Busawamanze HC III':299,'Buwambo HC IV':1017,
        'COMMUNITY HEALTH PLAN UGANDA':615,'Ggwatiro Nursing Home HC III':343,'Gombe (Wakiso) HC II':15,
         'Kabubbu HC II':673,'Kasangati HC IV':3052,'Kawanda HC III':981,'Kira HC III':1457,'Kiziba HC III':480,'Mende HC III':248,
         'Nabutiti HC III':199,'Nabweru HC III':1603,'Namayumba HC IV':1963,'Namulonge HC III':471,'Nansana HC II':118,
         'Nassolo Wamala HC II':182,'Triam Medical Centre CLINIC-NR':260,'Ttikalu HC III':433,'Wakiso Banda HC II GOVT':38,
         'Wakiso Epi Centre HC III GOVT':631,'Wakiso HC IV':3652,'Wakiso Kasozi HC III GOVT':233,'Watubba HC III':536,'Kakiri HC III':961}

KALANGALA= {'Bubeke HC III': 611,'Bufumira HC III': 405,'Bukasa HC IV': 1029, 'Bwendero HC III':1007,'Jaana HC II':13,
           'Kachanga Island HC II':219,'Kalangala HC IV':1443, 'Kasekulo HC II': 6,'Lujjabwa Island HC II': 345,'Lulamba HC III': 647,
           'Mazinga HC III': 524,'Mugoye HC III': 1131,'Mulabana HC II': 16,'Ssese Islands African Aids Project (SIAA)':20}  

LWENGO = {'Katovu HC III':454, 'Kiwangala HC IV': 1595, 
         'Kyazanga HC IV': 1962,'Kyetume HC III': 532, 'Lwengo HC IV': 1444, 'KINONI':2257,'Nanywa HC III':483,
         }

ENTEBBE = {'Bussi HC III': 219, 'Bweyogerere HC III': 966, 'BUNAMWAYA H-C II':32,'JCRC (Wakiso)':13247,'Kasenge H-C II':62, 'Kajjansi HC III':1928, 'Kasanje HC III': 826,
'Kigungu HC III':611, 'Kirinya H-C II':47, 
           'Kyengera HC III':578, 'Lufuka Valley HC III': 219, 'Mildmay Uganda HOSPITAL':14535, 'Mutundwe HC II':45,'Mutungo HC II':45, 'Nakawuka HC III':1092, 'Nalugala HC II':68,
           'Ndejje HC IV':1922, 'Nsangi HC III':2572, 'Seguku HC II':91, 'TASO Entebbe CLINIC' :6341, 'Wagagai HC IV': 531,'ZINGA HC II':246,'Kasoozo H-C III':30,'Katabi H-C III':85,
           'Kimwanyi H-C III':23, 'Kireka H-C II':52, 'KYENGEZA H-C II': 10, 'LUBBE H-C II':9, 'MAGANJO  H-C II':28, 'MAGOGGO H-C II': 16, 'Matugga H-C III':62,
           'Migadde H-C II':11, 'Namugongo Fund For Special Children': 589, 'NSAGGU H-C II':29, 'Nurture Africa H-C III':2369, 'Kitala HC II':175
}


districts = ['BUKOMANSIMBI', 'BUTAMBALA','ENTEBBE HUB', 'GOMBA','KALANGALA','KALUNGU', 'KYOTERA', 
             'LYANTONDE','LWENGO','MASAKA CITY', 'MASAKA DISTRICT', 'MPIGI', 'RAKAI', 'SEMBABULE', 'WAKISO HUB']
st.write('BEING UPDATED')
st.stop()

st.success('WELCOME, this app was developed by Dr. Luminsa Desire, for any concern, reach out to him at desireluminsa@gmail.com')
district = st.selectbox('Select a district:', districts, index=None)

file = st.file_uploader('Upload your CPHL extract here')


ext = None
if file is not None:
    # Get the file name
    fileN = file.name
    ext = os.path.splitext(fileN)[1]

df = None
if file and district is not None:
    if ext == '.csv':  # Compare with '.csv'
        df = pd.read_csv(file)
    else:
        st.write('This may not be a CPHL extract, it must be in CSV form.')
    # Display DataFrame
    if df is not None and district is not None:
        df['facility'] =  df['facility'].str.replace('/', '-')
        df['facility'] =  df['facility'].str.replace('Kinoni Welfare Medical Centre CLINIC', 'KINONI')
        df['facility'] =  df['facility'].str.replace('Mukwano Medical Centre CLINIC', 'Lukaya HC III')
        df['facility'] =  df['facility'].str.replace('St. Francis Maternity Home HC II', 'Lukaya HC III')
        df['facility'] =  df['facility'].str.replace('Teguzibirwa Dom Clinic', 'Lukaya HC III')
        if district  == 'BUKOMANSIMBI':
            fac = pd.DataFrame(list(BUKOMANSIMBI.items()), columns=['facility', 'Q2CURR'])
        elif district  == 'SEMBABULE':
            fac = pd.DataFrame(list(SEMBABULE.items()), columns=['facility', 'Q2CURR'])
        elif district  == 'MASAKA CITY':
            fac = pd.DataFrame(list(MASAKA_CITY.items()), columns=['facility', 'Q2CURR'])
        elif district  == 'MASAKA DISTRICT':
            fac = pd.DataFrame(list(MASAKA_DISTRICT.items()), columns=['facility', 'Q2CURR'])
        elif district  == 'KALUNGU':
            fac = pd.DataFrame(list(KALUNGU.items()), columns=['facility', 'Q2CURR'])
        elif district == 'MPIGI':
            fac = pd.DataFrame(list(MPIGI.items()), columns=['facility', 'Q2CURR'])
        elif district  == 'BUTAMBALA':
            fac = pd.DataFrame(list(BUTAMBALA.items()), columns=['facility', 'Q2CURR'])
        elif district  == 'GOMBA':
            fac = pd.DataFrame(list(GOMBA.items()), columns=['facility', 'Q2CURR'])
        elif district  == 'KYOTERA':
            fac = pd.DataFrame(list(KYOTERA.items()), columns=['facility', 'Q2CURR'])
        elif district  == 'RAKAI':
            fac = pd.DataFrame(list(RAKAI.items()), columns=['facility', 'Q2CURR'])
        elif district  == 'KALANGALA':
            fac = pd.DataFrame(list(KALANGALA.items()), columns=['facility', 'Q2CURR'])
        elif district  == 'LYANTONDE':
            fac = pd.DataFrame(list(LYANTONDE.items()), columns=['facility', 'Q2CURR'])
        elif district  == 'LWENGO':
            fac = pd.DataFrame(list(LWENGO.items()), columns=['facility', 'Q2CURR'])
        elif district == 'WAKISO HUB':
            fac = pd.DataFrame(list(WAKISO.items()), columns=['facility', 'Q2CURR'])
        elif district == 'ENTEBBE HUB':
            fac = pd.DataFrame(list(ENTEBBE.items()), columns=['facility', 'Q2CURR'])
        else:
            print('NO DISTRICT CHOSEN')

        coldist = fac['facility'].unique().tolist()
        colextr = df['facility'].unique().tolist()
        emrcolumns= ['A', 'RE', 'VOB']
        
        for column in coldist:
            if column not in colextr:
                st.write (f'**THIS EXTRACT DOES NOT HAVE FACILITIES IN {district}**')
                st.write('**You either uploaded a wrong exract or chose a wrong district, please try again!!**')
                st.stop()
            else:
                columns = fac['facility'].unique().tolist()
                df = df[df['facility'].isin(columns)].copy()
                df['ART-NUMERIC'] = df['art_number'].replace('[^0-9]','',regex=True)
                df['dCOL'] = df['date_collected'].astype(str)
                
                
                df['dCOL'] = df['dCOL'].str.replace('/', '*')
                df['dCOL'] = df['dCOL'].str.replace('-', '*')
                #df['dCOL'] = df['dCOL'].str.replace('/', '*')
                
                
                df[['Dyear', 'Dmonth', 'Dday']] = df['dCOL'].str.split('*', expand=True)
                
                df[['Dyear', 'Dmonth', 'Dday']]= df[['Dyear', 'Dmonth', 'Dday']].apply(pd.to_numeric, errors='coerce')
                 
                df['Dyear'] = df['Dyear'].fillna(2022)
                a = df[df['Dyear']>31].copy()
                b = df[df['Dyear']<32].copy()
                b = b.rename(columns={'Dyear': 'Dday1', 'Dday': 'Dyear'})
                b = b.rename(columns={'Dday1': 'Dday'})
                df = pd.concat([a,b])
                df[['Dyear', 'Dmonth', 'Dday']]= df[['Dyear', 'Dmonth', 'Dday']].apply(pd.to_numeric, errors='coerce')
                df = df[((df['Dyear']==2024) | ((df['Dyear']==2023) & (df['Dmonth']>9)))].copy()
                df = df.sort_values(by= ['Dyear', 'Dmonth', 'Dday'], ascending=False)

                def Viremia (x):
                    if 0<= x <= 200:
                        return 'Suppressed'
                    elif 201 <= x <= 399:
                        return 'LLV'
                    elif x == 400:
                        return 'suppressed'
                    elif 401 <= x <= 999:
                        return 'LLV'
                    elif x >= 1000:
                        return 'HLV'
                    else:
                        return None
                
                df['result_numeric'] = pd.to_numeric(df['result_numeric'],errors='coerce')
                df['SUP']= df['result_numeric'].apply(Viremia)
                facilities = df['facility'].unique()
                dfdups = df.copy()
                dfa = []
                for facility in facilities:
                    dfs = df[df['facility']==facility]
                    dfs = dfs.sort_values(by= ['Dyear', 'Dmonth', 'Dday'], ascending=False)
                    dfs['ART-NUMERIC'] =  pd.to_numeric(dfs['ART-NUMERIC'], errors='coerce') 
                    dfs = dfs.drop_duplicates(subset='ART-NUMERIC', keep='first')
                    dfs =dfs[['facility','ART-NUMERIC','art_number','date_collected','Dyear', 'Dmonth', 'Dday','result_numeric','SUP']]
                    name = f'{facility}'
                    dfa.append(dfs)
                dy = pd.concat(dfa) 
            
                dfnodups = dy.copy()  
                pivot = pd.pivot_table(dy, index='facility', values='ART-NUMERIC', aggfunc='count')
                dta = pivot.reset_index()
                dta = dta.rename(columns={'ART-NUMERIC':'BLEEDS'}) 
                dy['SUP'] = dy['SUP'].astype(str)
                NS = dy[(dy['SUP']== 'HLV') | (dy['SUP']=='LLV')].copy()
                NS[['Dyear', 'Dmonth']] = NS[['Dyear', 'Dmonth']].apply(pd.to_numeric, errors= 'coerce')
                NS = NS[((NS['Dyear']==2024)| ((NS['Dyear']==2023) & (NS['Dmonth']>9)))]
                HLV = NS[(NS['SUP']== 'HLV')].copy()
                LLV = NS[(NS['SUP']== 'LLV')].copy()
                pivo = pd.pivot_table(HLV, index='facility', values='ART-NUMERIC', aggfunc='count')
                dtb = pivo.reset_index()
                dtb = dtb.rename(columns={'ART-NUMERIC':'HLVs'})
                piv = pd.pivot_table(LLV, index='facility', values='ART-NUMERIC', aggfunc='count')
                dtc = piv.reset_index()
                dtc = dtc.rename(columns={'ART-NUMERIC':'LLVs'})
                dfa = pd.merge(fac,dta, on = 'facility', how = 'left')
                dfb = pd.merge(dfa,dtb, on = 'facility', how = 'left')
                dfc = pd.merge(dfb,dtc, on = 'facility', how = 'left')
            
                #file = r"C:\Users\Desire Lumisa\Desktop\New folder (2)\THISBP.csv"
                dfc[['Q2CURR', 'BLEEDS', 'HLVs', 'LLVs']] = dfc[['Q2CURR', 'BLEEDS', 'HLVs', 'LLVs']].apply(pd.to_numeric, errors='coerce')
                dfc['VL COV'] = (dfc['BLEEDS']*100)/ (dfc['Q2CURR'])
                dfc['VL COV'] = dfc['VL COV'].astype(int)
                dfc['BALANCE'] = (dfc['Q2CURR']*0.95)-(dfc['BLEEDS'])
                dfc['BALANCE'] = dfc['BALANCE'].astype(int)
                def achieve (v):
                    if v < 0:
                        return 0
                    else:
                        return v
                dfc['BALANCE TO 95%'] = dfc['BALANCE'].apply(achieve)
                dfc = dfc[['facility', 'Q2CURR', 'BLEEDS','VL COV','BALANCE TO 95%', 'HLVs', 'LLVs']]

                r = dfc['Q2CURR'].sum()
                t = dfc['BLEEDS'].sum()
                y = dfc['BALANCE TO 95%'].sum()
                u = dfc['HLVs'].sum()
                i = dfc['LLVs'].sum()
                o = int((t*100)/r)

                dfc.loc[len(dfc), 'facility'] = 'TOTAL'
                dfc.loc[len(dfc)-1, 'Q2CURR'] = r
                dfc.loc[len(dfc)-1, 'BLEEDS'] = t
                dfc.loc[len(dfc)-1, 'VL COV'] = o
                dfc.loc[len(dfc)-1, 'BALANCE TO 95%'] = y
                dfc.loc[len(dfc)-1, 'HLVs'] = u
                dfc.loc[len(dfc)-1, 'LLVs'] = i
    if df is not None:           
        dfe = dfc.set_index('facility')            
        st.write(dfe.head(2))     
    if df is not None:        
       # if st.button('DOWNLOAD FILE FOR VL COVERAGE ', key='active'):
                wb = Workbook()
                ws = wb.active
                # Convert DataFrame to Excel
                for r_idx, row in enumerate(dfc.iterrows(), start=1):
                    for c_idx, value in enumerate(row[1], start=1):
                                ws.cell(row=r_idx, column=c_idx, value=value)

                ws.insert_rows(0)
                ws['A1'] = 'FACILITY'
                ws['B1'] = 'Q2 CURR'
                ws['C1'] = 'BLEEDS'
                ws['D1'] = 'VL COV'
                ws['E1'] = 'BALANCE TO 95%'
                ws['F1'] = 'HLVs'
                ws['G1'] = 'LLVs'
                
                max_row = ws.max_row
                ws.cell(row=max_row, column=1).alignment = Alignment(horizontal = 'center')
    
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['E'].width = 17
                ws.column_dimensions['B'].width = 10
                ws.column_dimensions['D'].width = 10

                #letters = ['A1', 'B1', 'C1', 'D1']
                

                letter = 'D'
                red = PatternFill(fill_type = 'solid', start_color = 'ff0000')
                yellow = PatternFill(fill_type = 'solid', start_color = 'ffff00')
                green = PatternFill(fill_type = 'solid', start_color = '04AA6D')

                for num in range(2, ws.max_row +1):
                    ws[f'{letter}{num}'].alignment = Alignment(horizontal='center')
                    if ws[f'{letter}{num}'].value <85:
                        ws[f'{letter}{num}'].fill = red
                    elif ws[f'{letter}{num}'].value <95:
                        ws[f'{letter}{num}'].fill = yellow   
                    else:
                        ws[f'{letter}{num}'].fill = green
                        ws[f'{letter}{num}'].border = Border(top= Side(style = 'thick'),
                                                        left= Side(style = 'thick'),
                                                        right= Side(style = 'thick'),
                                                        bottom= Side(style = 'thick')) 
                    
                blue = PatternFill(fill_type = 'solid', start_color = '80F5F5')
                letter = ['A1', 'B1', 'C1', 'D1','E1','F1','G1']
                for each in letter:
                    ws[f'{each}'].font = Font(b= True, i = True)
                    ws[f'{each}'].fill = blue
                    ws[f'{each}'].border = Border(top = Side(style = 'thin', color ='000000'),
                                                            right = Side(style = 'thin', color ='000000'),
                                                            left = Side(style = 'thin', color ='000000'),
                                                            bottom = Side(style = 'thin', color ='000000'))
              
                grey = PatternFill(fill_type = 'solid', start_color = 'ECF1F1')
                letter = ['A', 'B', 'C','E','F','G']
                for each in letter:
                    ws[f'{each}{max_row}'].font = Font(b= True, i = True)
                    ws[f'{each}{max_row}'].fill = grey
                    ws[f'{each}{max_row}'].border = Border(top = Side(style = 'thin', color ='000000'),
                                                            right = Side(style = 'thin', color ='000000'),
                                                            left = Side(style = 'thin', color ='000000'),
                                                            bottom = Side(style = 'thin', color ='000000'))



                ws.sheet_view.ShowGridLines = False        

                ran = random.random()
                rand = round(ran,2)
                file_path = os.path.join(os.path.expanduser('~'), 'Downloads', f'{district}VL_COV {rand}.xlsx')
                directory = os.path.dirname(file_path)
                Path(directory).mkdir(parents=True, exist_ok=True)

                  # Save the workbook
                wb.save(file_path)
                # Serve the file for download
                with open(file_path, 'rb') as f:
                      file_contents = f.read()           
                st.download_button(label=f'DONLOAD VL COV FOR {district} ', data=file_contents,file_name=f' {district} VL COV {rand}.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
             
            
    if df is not None and district is not None:
        def download_with_duplicates(df):
            st.write(f"<h6>CSV FILES for {district} WITH NO DUPLICATES</h6>", unsafe_allow_html=True)

            if df is not None and district is not None:
                dft = dfnodups.copy()
                uniques = dft['facility'].unique()

                # Create an expander to contain the download buttons
                with st.expander(f"Download files for {district} Facilities without duplicates"):
                    for facility in uniques:
                        dfs = dft[dft['facility'] == facility]
                        dfs = dfs[['facility', 'ART-NUMERIC', 'art_number', 'date_collected', 'Dyear', 'Dmonth', 'Dday', 'result_numeric']]
                        csv_data = dfs.to_csv(index=False)

                        # Create a download button for each facility
                        st.download_button(
                            label=f"Download CSV for {facility} without duplicates",
                            data=csv_data,
                            file_name=f"{facility}_data_without_duplicates.csv",
                            mime="text/csv"
                        )

        def download_without_duplicates(df):
            st.write(f"<h6>CSV FILES for {district} WITH DUPLICATES</h6>", unsafe_allow_html=True)

            if df is not None and district is not None:
                dft = dfdups.copy()
                uniques = dft['facility'].unique()

                # Create an expander to contain the download buttons
                with st.expander(f"Download files for {district} Facilities with duplicates)"):
                    for facility in uniques:
                        dfs = dft[dft['facility'] == facility]
                        dfs = dfs[['facility', 'ART-NUMERIC', 'art_number', 'date_collected', 'Dyear', 'Dmonth', 'Dday', 'result_numeric']]
                        csv_data = dfs.to_csv(index=False)

                        # Create a download button for each facility
                        st.download_button(
                            label=f"Download CSV for {facility} with duplicates",
                            data=csv_data,
                            file_name=f"{facility}_data_with_duplicates.csv",
                            mime="text/csv"
                        )

        def main():
            # Call the download functions
            download_with_duplicates(df)
            download_without_duplicates(df)

        if __name__ == "__main__":
            main()








