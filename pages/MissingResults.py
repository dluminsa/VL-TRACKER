import streamlit as st 
import pandas as pd
import os
import glob
import numpy as np
import random
from openpyxl import Workbook
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
# from openpyxl import * #load_workbook
# from openpyxl.styles import *
#import numpy as np
cola, colb = st.columns([1,1])
st.write('BEING UPDATED')
st.stop()
st.success('WELCOME, this app was developed by Dr. Luminsa Desire, for any concern, reach out to him at desireluminsa@gmail.com')
st.markdown('**Follow these simple instructions before you proceed:**')
cola,colb = st.columns([1,2])
cola.markdown('**You need two extracts; *a CPHL extract and an emr extract* for this facility**')
colb.markdown('**In the emr extract:**')
colb.markdown('Rename the **HIV Clinic NO.** column to **A**')
#colb.markdown('Rename the **VIral load results** column to **RE**')
colb.markdown('Rename the **HIV VIRAL LOAD DATE** column to **VD**')
cola.markdown('Rename the **RETURN VISIT DATE** column to **RD**')
 
col1,col2 = st.columns([1,1])
cphl = col1.file_uploader(label='Upload your CPHL extract here')
col1.write('The cphl extract must be in csv form')
emr = col2.file_uploader(label='Upload your EMR extract here')
col2.write('The emr extract must be in xlsx form')

extcphl = None
extemr = None
if cphl is not None and emr is not None:
    # Get the file name
    filecphl = cphl.name
    extcphl = os.path.splitext(filecphl)[1]
    fileemr = emr.name
    extemr = os.path.splitext(fileemr)[1]

df =None
df2 = None

df = None
if cphl is not None and emr is not None:
    if extcphl !='.csv':  # Compare with '.csv'
        st.write('Use a csv file generated from CPHL')
        st.stop()
    else:
        df = pd.read_csv(cphl)
        if extemr != '.xlsx': 
            st.write('First save the emr extract as an xlsx file')
            st.stop()
        else:
            dfc = pd.read_excel(emr)
            cphlcolumns = ['facility', 'ART-NUMERIC', 'art_number', 'date_collected', 'Dyear',
        'Dmonth', 'Dday', 'result_numeric', ]
            colcphl = df.columns.to_list()
            for column in cphlcolumns:
                if column not in colcphl:
                    st.write (f' ERROR !!! {column} is not in this CPHL extract')
                    st.stop()
                #print( 'REACH OUT TO YOUR TEAM LEAD FOR THE RIGHT EXTRACT')
            #print('kindly check the table above to see how to rename the columns')
            emrcolumns= ['A', 'RD','VD']
            colemr = dfc.columns.to_list()
            for column in emrcolumns:
                if column not in colemr:
                    st.write (f' ERROR !!! {column} is not in this EMR extract')
                    st.stop()
                else:
                    df[['ART-NUMERIC', 'Dyear','Dmonth', 'Dday']] = df[['ART-NUMERIC', 'Dyear','Dmonth', 'Dday']].apply(pd.to_numeric, errors='coerce')
                    df = df[((df['Dyear']==2024) | ((df['Dyear']==2023) & (df['Dmonth']>6)))].copy()
                    named = df['facility'].unique()
                    named = np.array2string(named)
                    named = named.strip('[]')
                    dfc['ART'] = dfc['A'].replace('[^0-9]','', regex = True)
                    dfc = dfc.dropna(subset=['ART'])
                    dfc['VOD'] = dfc['VD']
                    dfc['RT'] = dfc['RD']
                    y = pd.DataFrame({'RT':['1-1-50',1,'1/1/50'],'VOD':['1-1-50',1,'1/1/50'] })  
                    dfc = pd.concat([dfc,y])
                   
                  
                #THE VOD DATE
                    dfc['VOD'] = dfc['VOD'].astype(str)
                    dfc['VOD'] = dfc['VOD'].str.replace('00:00:00', '')
                
                    A = dfc[dfc['VOD'].str.contains('-')]
                    a = dfc[~dfc['VOD'].str.contains('-')]
                    B = a[a['VOD'].str.contains('/')]
                    C = a[~a['VOD'].str.contains('/')]

                    A[['VOyear', 'VOmonth', 'VOday']] = A['VOD'].str.split('-', expand = True)
                    B[['VOyear', 'VOmonth', 'VOday']] = B['VOD'].str.split('/', expand = True)
                                
                    C['VOD'] = pd.to_numeric(C['VOD'], errors='coerce')
                    C['VOD'] = pd.to_datetime(C['VOD'], origin='1899-12-30', unit='D', errors='coerce')
                    C['VOD'] =  C['VOD'].astype(str)
                    C[['VOyear', 'VOmonth', 'VOday']] = C['VOD'].str.split('-', expand = True)
                    dfc = pd.concat([A,B,C])

                    #THE RT DATE
                    dfc['RT'] = dfc['RT'].astype(str)
                    dfc['RT'] = dfc['RT'].str.replace('00:00:00', '')
                    
                    
                    A = dfc[dfc['RT'].str.contains('-')]
                    a = dfc[~dfc['RT'].str.contains('-')]
                    B = a[a['RT'].str.contains('/')]
                    C = a[~a['RT'].str.contains('/')]

                    A[['Ryear', 'Rmonth', 'Rday']] = A['RT'].str.split('-', expand = True)
                    B[['Ryear', 'Rmonth', 'Rday']] = B['RT'].str.split('/', expand = True)
                                
                    C['RT'] = pd.to_numeric(C['RT'], errors='coerce')
                    C['RT'] = pd.to_datetime(C['RT'], origin='1899-12-30', unit='D', errors='coerce')
                    C['RT'] =  C['RT'].astype(str)
                    C[['Ryear', 'Rmonth', 'Rday']] = C['RT'].str.split('-', expand = True)
                    dfc = pd.concat([A,B,C])

                    dfc['RD'] = dfc['RD'].astype(str)
                    dfc['VD'] = dfc['VD'].astype(str)
    
                    dfc['RD'] = dfc['RD'].str.replace('00:00:00', '')
                    dfc['VD'] = dfc['VD'].str.replace('00:00:00', '')

                    dfc['Rday1'] = dfc['Rday'].astype(str).str.split('.').str[0]
                    dfc['Rmonth1'] = dfc['Rmonth'].astype(str).str.split('.').str[0]
                    dfc['Ryear1'] = dfc['Ryear'].astype(str).str.split('.').str[0]

                    dfc['Vday1'] = dfc['VOday'].astype(str).str.split('.').str[0]
                    dfc['Vmonth1'] = dfc['VOmonth'].astype(str).str.split('.').str[0]
                    dfc['Vyear1'] = dfc['VOyear'].astype(str).str.split('.').str[0]
                 
                    dfc['RETURN DATE'] = dfc['Rday1'] + '/' + dfc['Rmonth1'] + '/' + dfc['Ryear1']
                    dfc['VL DATE'] = dfc['Vday1'] + '/' + dfc['Vmonth1'] + '/' + dfc['Vyear1']

                    dfc['RETURN DATE'] = pd.to_datetime(dfc['RETURN DATE'], format='%d/%m/%Y', errors='coerce')
                    dfc['VL DATE'] = pd.to_datetime(dfc['VL DATE'], format='%d/%m/%Y', errors='coerce')

                    dfc['RETURN DATE'] = dfc['RETURN DATE'].astype(str)
                    dfc['VL DATE'] = dfc['VL DATE'].astype(str)
                    dfc['RETURN DATE'] = dfc['RETURN DATE'].str.replace('00:00:00', '')
                    dfc['VL DATE'] = dfc['VL DATE'].str.replace('00:00:00', '')
         
                    dfc[['VOyear', 'VOmonth', 'VOday']] =dfc[['VOyear', 'VOmonth', 'VOday']].apply(pd.to_numeric, errors = 'coerce')
                    dfc[['Ryear', 'Rmonth', 'Rday']] =dfc[['Ryear', 'Rmonth', 'Rday']].apply(pd.to_numeric, errors = 'coerce')
                   
                    dfc['Ryear'] = dfc['Ryear'].fillna(2022)
                    dfc['VOyear'] = dfc['VOyear'].fillna(2022)
                    #st.write(dfc.columns)
                    #st.stop()
                    dfc['Ryear'] = dfc['Ryear'].astype(int)

                    e = dfc[dfc['Ryear']>31].copy()     
                    f = dfc[dfc['Ryear']<32].copy()
                    #st.stop()
                    f = f.rename(columns={'Ryear': 'Rday2', 'Rday': 'Ryear'})
                    f = f.rename(columns={'Rday2': 'Rday'})
                    dfc = pd.concat([e,f])
                    #st.write(dfc.head(5))
                    #st.write(dfc.columns)
                    #st.write( dfc['VOyear'])
                    #st.stop()
                    a = dfc[dfc['VOyear']>31].copy()
                    b = dfc[dfc['VOyear']<32].copy()
                    b = b.rename(columns={'VOyear': 'VOday2', 'VOday': 'VOyear'})
                    b = b.rename(columns={'VOday2': 'VOday'})
                    dfc = pd.concat([a,b])
                    def NEW (x, y):
                        if x ==2024:
                                return 'NEW'
                        elif x == 2023 and y > 5:
                            return 'NEW'
                        else:
                            return 'OLD'  
                    dfc['RESULTS'] = dfc.apply(lambda row: NEW(row['VOyear'], row['VOmonth']), axis=1)
                    dfc[['Ryear', 'Rmonth', 'Rday']] =dfc[['Ryear', 'Rmonth', 'Rday']].apply(pd.to_numeric, errors = 'coerce')
                    dfd = dfc[dfc['Ryear']>2024].copy()
                    dfe = dfc[dfc['Ryear']==2024].copy()
                    dfe = dfe[((dfe['Rmonth']>3) | ((dfe['Rmonth']==3) & (dfe['Rday']>3)))].copy()
                    dfc = pd.concat([dfd,dfe])
                 
                    dfj = dfc[dfc['RESULTS']== 'OLD']
                    dfj = dfj[['ART','A', 'RD', 'VD','VL DATE', 'RETURN DATE']]
                    dfj['ART'] = pd.to_numeric(dfj['ART'])
                    df[ 'ART-NUMERIC'] = pd.to_numeric(df[ 'ART-NUMERIC'])
                    dfk = dfj[dfj['ART'].isin(df[ 'ART-NUMERIC'])]
                    dfh = df[df['ART-NUMERIC'].isin(dfj[ 'ART'])]
                    dfh = dfh.rename(columns={'ART-NUMERIC':'ART'})
                    dft = pd.merge(dfk, dfh, on = 'ART', how= 'left')
                 
                    dft = dft.rename(columns = {'A':'ART-NO'})
                    dft['date_collected'] =  dft['date_collected'].astype(str)
                    dft['date_collected'] =  dft['date_collected'].str.replace('*', '-')
                    dft = dft[['ART-NO', 'RETURN DATE', 'VL DATE','art_number','date_collected','result_numeric']]
                    
    #if df is not None and df2 is not None: 
    a = dft.shape[0]
    st.success(f'I see over **{a}** results at CPHL that are not yet entered into EMR')
                    


    # st.success('Analysis done')
    #if st.button('DOWNLOAD FILE'):
    wb = Workbook()
    ws = wb.active
 
    # Convert DataFrame to Excel
    for r_idx, row in enumerate(dft.iterrows(), start=1):
           for c_idx, value in enumerate(row[1], start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    ws.insert_rows(0,2)

    blue = PatternFill(fill_type = 'solid', start_color = 'F6F8F7')
        # ws.column_dimensions['H'].width = 14

    for num in range (1, ws.max_row+1):
         for letter in ['D','E', 'F']:
              ws[f'{letter}{num}'].font = Font(b= True, i = True)
              ws[f'{letter}{num}'].font = Font(b= True, i = True)
              ws[f'{letter}{num}'].fill = blue
              ws[f'{letter}{num}'].border = Border(top = Side(style = 'thin', color ='000000'),
                                                    right = Side(style = 'thin', color ='000000'),
                                                    left = Side(style = 'thin', color ='000000'),
                                                    bottom = Side(style = 'thin', color ='000000'))
    ws['B1'] ='EMR DETAILS'
    ws['F1'] = 'CPHL DETAILS'
    ws['A2'] = 'ART-NO'
    ws['B2'] = 'RETUR VISIT DATE'
    ws['C2'] = 'EMR VL DATE' 
    ws['D2'] = 'ART NO'
    ws['E2'] = 'CPHL DATE'
    ws['F2']  = 'CPHL RESULTS'


    letters = ['B', 'C', 'D','E','F']
    for letter in letters:
          ws.column_dimensions[letter].width =15

    ran = random.random()
    rand = round(ran,2)
    file_path = os.path.join(os.path.expanduser('~'), 'Downloads', f'{named}_missing_results {rand}.xlsx')
    directory = os.path.dirname(file_path)
    Path(directory).mkdir(parents=True, exist_ok=True)

                  # Save the workbook
    wb.save(file_path)
     # Serve the file for download
    with open(file_path, 'rb') as f:
              file_contents = f.read()           
    st.download_button(label=f'DONLOAD MISSING RESULTS FOR {named} ', data=file_contents,file_name=f'{named}_missing_results {rand}.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

       
    if df is not None:
        dfu = dft.set_index('ART-NO')
        st.write(dfu.head(3))


